# -*- coding: utf-8 -*-
"""
SFMC → AJO HTML translator with structured Excel logging.

Key points:
- Converts AMPScript IF/ELSE blocks to AJO Liquid ({% if %}...{%/if%}).
- Replaces SFMC print tokens (%%=v(@var)=%%) to Liquid ({{ expr }}) ONLY when
  the variable has a clear mapping from the Excel de-para sheet.
- Handles the "the + PlanLegalName" grammar logic inside the THEN branch
  to keep natural language intact.
- Safely comments any remaining AMPScript, including tokens embedded in HTML
  attributes (e.g., src="%%=...=%%") without breaking the markup.
- Hoists any `{% let ... %}` found inside AMPScript blocks so they remain active.
- Emits a neat Excel log with four tabs: Summary, Variables, Unmapped_Variables,
  and Commented_AMPScript.

Author: (team)
"""

import re
import sys
import pandas as pd
from pathlib import Path
from datetime import datetime

# --- Excel writer (openpyxl) -----------------------------------------------
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    raise SystemExit(
        "\nMissing dependency: 'openpyxl'\n"
        "Install it with:  py -m pip install openpyxl\n"
    )

# ==============================
# Configuration (CLI defaults)
# ==============================
html_input  = sys.argv[1] if len(sys.argv) > 1 else "input.html"
html_output = sys.argv[2] if len(sys.argv) > 2 else "output.html"
excel_path  = sys.argv[3] if len(sys.argv) > 3 else "SFMCtoAJOComparision.xlsx"


# ==============================
# Small helpers
# ==============================
def pick_cols(df: pd.DataFrame):
    """
    Resolve SFMC/AJO columns even if headers change slightly.
    """
    cols = {str(c).strip().lower(): c for c in df.columns}
    sfmc_col = next((v for k, v in cols.items() if "sfmc" in k), df.columns[0])
    ajo_col  = next((v for k, v in cols.items() if "ajo"  in k), df.columns[1])
    return sfmc_col, ajo_col


def build_flex_regex(sfmc_snippet: str) -> re.Pattern:
    """
    Build a forgiving regex for de-para replacements:
    - Ignore whitespace differences,
    - Allow spaces around '='.
    """
    s = re.escape(str(sfmc_snippet).strip())
    s = re.sub(r'\\[ \t\r\n\f]+', r'\\s*', s)
    s = s.replace(r'\=', r'\s*=\s*')
    return re.compile(s, flags=re.IGNORECASE | re.DOTALL)


def month_abbr_en(dt: datetime) -> str:
    """Return English month abbreviation regardless of OS locale."""
    return ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"][dt.month-1]


def make_log_filename(dt: datetime) -> str:
    """Build the requested Excel log filename format."""
    return f"Output_Log_{dt.year}_{month_abbr_en(dt)}_{dt.day:02d}_{dt:%H}h{dt:%M}min.xlsx"


# ===========================================
# Mapping extraction from the Excel de-para
# ===========================================
def extract_var_expressions(df: pd.DataFrame):
    """
    Learn mappings for SFMC variables from the de-para sheet.

    Rules:
    - Only trust a row if the SFMC side references exactly ONE variable (@var).
      This avoids polluting @the with PlanLegalName, etc.
    - Skip helper words like @the entirely (reserved).
    - If AJO cell contains:
        * {% let Var = ... %}  → use "Var" as the expression (variable name).
        * {{ expr }}           → use inner expr.
        * profile./context.    → use that path directly.
    - Also track 'covered' variables to avoid false "unmapped" warnings.
    """
    var_expr = {}
    covered = set()

    let_re   = re.compile(r'{%\s*let\s+([A-Za-z_]\w*)\s*=\s*\(?\s*(.*?)\s*\)?\s*%}', re.IGNORECASE | re.DOTALL)
    print_re = re.compile(r'{{\s*(.*?)\s*}}', re.DOTALL)
    path_re  = re.compile(r'\b(?:profile|context)\.[A-Za-z0-9_\.\[\]]+', re.IGNORECASE)

    RESERVED_SKIP = {"the"}  # never auto-map helper words like @the

    for _, row in df.iterrows():
        sfmc = str(row["SFMC"] or "")
        ajo  = str(row["AJO"] or "")

        vars_in_sfmc = re.findall(r'@(\w+)', sfmc)
        if not vars_in_sfmc:
            continue

        # Only accept if the row mentions exactly one unique variable
        unique_vars = list({v.lower() for v in vars_in_sfmc})
        single_var = (len(unique_vars) == 1)

        expr = None
        m_let = let_re.search(ajo)
        if m_let:
            expr = m_let.group(1).strip()
        else:
            m_print = print_re.search(ajo)
            if m_print:
                expr = m_print.group(1).strip()
            else:
                m_path = path_re.search(ajo)
                if m_path:
                    expr = m_path.group(0).strip()

        for var in vars_in_sfmc:
            vlow = var.lower()
            if vlow in RESERVED_SKIP:
                continue
            if expr and single_var:
                var_expr[vlow] = expr
                covered.add(vlow)
            elif any(x in ajo for x in ["profile.", "context.", "fragment", "{{", "{%"]):
                # AJO side suggests a destination; don't warn about it even if we didn't extract a clean name.
                covered.add(vlow)

    return var_expr, covered


def resolve_expr(var_name: str, var_expr: dict) -> str:
    """
    Return the Liquid expression to use for @var.
    If there's no mapping, return the raw name so we know NOT to convert it.
    """
    return var_expr.get(var_name.lower(), var_name)


# =====================================
# AMPScript condition → Liquid (AJO)
# =====================================
def translate_condition_to_liquid(cond: str, warnings: list, var_expr: dict, covered_vars: set) -> str:
    """
    Translate common AMPScript conditional patterns to AJO Liquid:
    - NOT EMPTY(@x)      → {% if length(expr) > 0 %}
    - EMPTY(@x)          → {% if length(expr) == 0 %}
    - CONTAINS(@x,'v')   → {% if contains(expr, 'v') %}
    - @x == 'v' / != 'v' → {% if expr == 'v' %} / {% if expr != 'v' %}
    - @x (truthy)        → {% if expr %}
    Unknown patterns are left as an HTML comment so review is easy.
    """
    cond_original = cond
    cond_l = cond.strip()

    # Keep track of variables referenced within the condition to report unmapped ones later
    for v in re.findall(r'@(\w+)', cond_l, flags=re.IGNORECASE):
        if v.lower() not in covered_vars and v.lower() not in {"email","firstname","lastname","country"}:
            warnings.append(v)

    re_not_empty = re.compile(r'not\s*empty\s*\(\s*@(\w+)\s*\)', re.IGNORECASE)
    re_empty     = re.compile(r'empty\s*\(\s*@(\w+)\s*\)', re.IGNORECASE)
    re_contains  = re.compile(r'contains\s*\(\s*@(\w+)\s*,\s*[\'"]([^\'"]+)[\'"]\s*\)', re.IGNORECASE)
    re_cmp       = re.compile(r'@(\w+)\s*(==|!=)\s*[\'"]([^\'"]+)[\'"]', re.IGNORECASE)
    re_truthy    = re.compile(r'^\s*@(\w+)\s*$', re.IGNORECASE)

    def learn_expr(var_name: str) -> str:
        expr = resolve_expr(var_name, var_expr)
        if expr != var_name:
            covered_vars.add(var_name.lower())
        return expr

    m = re_not_empty.search(cond_l)
    if m:
        e = learn_expr(m.group(1))
        return "{% if length(" + e + ") > 0 %}"

    m = re_empty.search(cond_l)
    if m:
        e = learn_expr(m.group(1))
        return "{% if length(" + e + ") == 0 %}"

    m = re_contains.search(cond_l)
    if m:
        e = learn_expr(m.group(1))
        val = m.group(2)
        return "{% if contains(" + e + ", '" + val + "') %}"

    m = re_cmp.search(cond_l)
    if m:
        e = learn_expr(m.group(1))
        op = "==" if m.group(2) == "==" else "!="
        val = m.group(3)
        return "{% if " + e + " " + op + " '" + val + "' %}"

    m = re_truthy.search(cond_l)
    if m:
        e = learn_expr(m.group(1))
        return "{% if " + e + " %}"

    # ELSEIF fallback: translate inner, then switch "if" to "elseif"
    if cond_l.lower().startswith("elseif "):
        inner = cond_l[7:].strip()
        trans = translate_condition_to_liquid(inner, warnings, var_expr, covered_vars)
        return trans.replace("{% if", "{% elseif")

    return "<!-- Untranslated condition: " + cond_original + " -->"


# ============================================
# Convert AMPScript IF/ELSE/ENDIF blocks
# ============================================
def convert_ampscript_if_blocks(html_text: str, warnings: list, var_expr: dict, covered_vars: set) -> str:
    """
    Convert blocks like:
      %%[ IF ... THEN ]%% ... %%[ ELSE ]%% ... %%[ ENDIF ]%%
      %%[ IF ... ]%%      ... %%[ ELSE ]%% ... %%[ ENDIF ]%%

    Also handles the special "the + PlanLegalName" grammar case within THEN:
    - Replaces "%%=v(@the)=%%%%=v(@planName)=%%" with:
        {% if PlanLegalName startsWith 'THE' %}{{ PlanLegalName }}{% else %}the {{ PlanLegalName }}{%/if%}
    """
    def repl(match):
        cond = (match.group(1) or "").strip()
        then_block = match.group(2) or ""
        else_block = match.group(4) or ""

        cond_liquid = translate_condition_to_liquid(cond, warnings, var_expr, covered_vars)

        # "the + PlanLegalName" case (doesn't depend on de-para)
        if re.search(r'%%=v\(@the\)=%%', then_block, flags=re.IGNORECASE) and \
           re.search(r'%%=v\(@planname\)=%%', then_block, flags=re.IGNORECASE):
            the_plan = (
                "{% if PlanLegalName startsWith 'THE' %}"
                "{{ PlanLegalName }}"
                "{% else %}the {{ PlanLegalName }}{%/if%}"
            )
            return cond_liquid + the_plan + ("{% else %}"+else_block if else_block.strip() else "") + "{%/if%}"

        # Convert prints ONLY when mapped in de-para
        def _print_sub(mv):
            var = mv.group(1)
            expr = resolve_expr(var, var_expr)
            if expr != var:
                return "{{ " + expr + " }}"
            return mv.group(0)  # keep AMPScript; it will be commented and logged later

        then_block = re.sub(r'%%=v\(@(\w+)\)=%%', _print_sub, then_block, flags=re.IGNORECASE)
        else_block = re.sub(r'%%=v\(@(\w+)\)=%%', _print_sub, else_block, flags=re.IGNORECASE)

        if else_block.strip():
            return cond_liquid + then_block + "{% else %}" + else_block + "{%/if%}"
        else:
            return cond_liquid + then_block + "{%/if%}"

    pattern = re.compile(
        r'%%\[\s*IF\s+(.+?)(?:\s+THEN)?\s*\]%%(.*?)((?:%%\[\s*ELSE\s*\]%%(.*?))?)%%\[\s*ENDIF\s*\]%%',
        re.IGNORECASE | re.DOTALL
    )
    return pattern.sub(repl, html_text)


# =========================================
# Global print replacement (only mapped)
# =========================================
def replace_all_prints(html_text: str, var_expr: dict) -> str:
    """
    Convert %%=v(@var)=%% → {{ expr }} only if @var has a known mapping.
    Otherwise keep AMPScript so the commenter can log it.
    """
    def _print_sub(mv):
        var = mv.group(1)
        expr = resolve_expr(var, var_expr)
        if expr != var:
            return "{{ " + expr + " }}"
        return mv.group(0)

    return re.sub(r'%%=v\(@(\w+)\)=%%', _print_sub, html_text, flags=re.IGNORECASE)


# =============================================================
# Comment AMPScript safely (handles attributes) + hoist {% let %}
# =============================================================
def comment_ampscript_with_hoist(html_text: str):
    """
    Two-phase approach to avoid breaking HTML:

    Phase 1 (attribute-safe):
      Replace any attribute value equal to an AMPScript token:
        attr="%%=...=%%"  →  attr="" <!-- %%=...=%% -->
      This keeps HTML valid and still surfaces the token for logging.

    Phase 2:
      For remaining AMPScript:
      - Hoist any `{% let ... %}` found inside AMPScript blocks so they remain active.
      - Comment the rest of the AMPScript:
          %%=...=%%, %%[ ... ]%%, <script runat=server>...</script>

    Returns:
      (new_html, commented_items)
      where commented_items is a list of (line_number, snippet) for the Excel log.
    """
    # --- Phase 1: neutralize tokens used as attribute values
    attr_token_re = re.compile(r'(\s[\w:-]+\s*=\s*)(["\'])(%%=.*?=%%)\2', re.DOTALL)
    commented = []

    def _line_of(idx: int) -> int:
        return html_text.count("\n", 0, idx) + 1

    parts = []
    last = 0
    for m in attr_token_re.finditer(html_text):
        start, end = m.span()
        parts.append(html_text[last:start])
        attr_prefix, quote, token = m.group(1), m.group(2), m.group(3)
        # Empty the attribute and place the token as an HTML comment right after it
        parts.append(attr_prefix + quote + quote + " <!-- " + token + " -->")
        commented.append((_line_of(start), token))
        last = end
    parts.append(html_text[last:])
    html_text = "".join(parts)

    # --- Phase 2: hoist LETs and comment remaining AMPScript
    block_re = re.compile(
        r'(%%=.+?=%%|%%\[[\s\S]*?\]%%|<script[^>]*\brunat=[\'"]server[\'"][^>]*>[\s\S]*?<\/script>)',
        re.IGNORECASE | re.DOTALL
    )
    let_re = re.compile(r'{%\s*let\s+[^%]+%}', re.IGNORECASE | re.DOTALL)

    # Precompute line offsets for context/line reporting
    lines = html_text.splitlines()
    offsets = [0]
    for line in lines:
        offsets.append(offsets[-1] + len(line) + 1)

    def find_line(pos: int) -> int:
        for i, end in enumerate(offsets):
            if pos < end:
                return max(1, i)
        return len(lines)

    out = []
    last_end = 0
    for m in block_re.finditer(html_text):
        start, end = m.span()
        full = m.group(1)
        line_no = find_line(start)

        out.append(html_text[last_end:start])

        if full.startswith("%%="):
            # Tokens inside attributes were already neutralized in Phase 1.
            # Any remaining token can be safely commented inline.
            out.append("<!-- " + full + " -->")
            commented.append((line_no, full))
            last_end = end
            continue

        if full.lower().startswith("%%[") or full.lower().startswith("<script"):
            # Hoist {% let ... %} and comment the rest
            lets = let_re.findall(full)
            cleaned = let_re.sub("", full)
            for l in lets:
                out.append(l)  # keep active
            if cleaned.strip():
                out.append("<!-- " + cleaned + " -->")
                commented.append((line_no, cleaned.strip()))
            last_end = end
            continue

        # Fallback (shouldn't happen): comment
        out.append("<!-- " + full + " -->")
        commented.append((line_no, full))
        last_end = end

    out.append(html_text[last_end:])
    return "".join(out), commented


# ==============================
# Excel logging
# ==============================
def write_excel_log(commented_blocks, total_found, total_replaced, output_path, warnings, covered_vars, var_expr):
    """
    Create a well-structured Excel logfile with:
    - Summary
    - Variables (learned from de-para)
    - Unmapped_Variables
    - Commented_AMPScript
    """
    now = datetime.now()
    xlsx_name = make_log_filename(now)
    wb = Workbook()

    # Styles
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="4F81BD")
    left     = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center   = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D9D9D9")

    def style_header(row):
        for c in row:
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = center
            c.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    def style_cells(ws):
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = left
                c.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Key", "Value"])
    style_header(ws[1])
    rows = [
        ("Run datetime", now.isoformat(timespec="seconds")),
        ("Input HTML", html_input),
        ("Output HTML", output_path),
        ("Mapping Excel", excel_path),
        ("SFMC matches found", total_found),
        ("Substitutions (table)", total_replaced),
        ("AMPScript blocks commented", len(commented_blocks)),
    ]
    for k, v in rows:
        ws.append([k, v])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 120
    style_cells(ws)
    ws.freeze_panes = "A2"

    # Variables sheet
    ws2 = wb.create_sheet("Variables")
    ws2.append(["SFMC @var", "AJO expression used"])
    style_header(ws2[1])
    for vlow, expr in sorted(var_expr.items()):
        ws2.append([f"@{vlow}", expr])
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 120
    style_cells(ws2)
    ws2.freeze_panes = "A2"

    # Unmapped variables sheet
    unresolved = sorted({w.lower() for w in warnings} - {v.lower() for v in covered_vars})
    ws3 = wb.create_sheet("Unmapped_Variables")
    ws3.append(["SFMC @var", "Action"])
    style_header(ws3[1])
    for name in unresolved:
        ws3.append([f"@{name}", f"Define profile.{name} or context.{name} in AJO (or add to mapping sheet)"])
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 120
    style_cells(ws3)
    ws3.freeze_panes = "A2"

    # Commented AMPScript sheet
    ws4 = wb.create_sheet("Commented_AMPScript")
    ws4.append(["Line", "Snippet"])
    style_header(ws4[1])
    for line, snippet in commented_blocks:
        ws4.append([line, snippet])
    ws4.column_dimensions["A"].width = 10
    ws4.column_dimensions["B"].width = 150  # wide for better readability
    style_cells(ws4)
    ws4.freeze_panes = "A2"

    wb.save(xlsx_name)
    return Path(xlsx_name)


# ==============================
# Pipeline
# ==============================
# 0) Read mapping sheet
df = pd.read_excel(excel_path, usecols="A:B")
sfmc_col, ajo_col = pick_cols(df)
df = df[[sfmc_col, ajo_col]].rename(columns={sfmc_col: "SFMC", ajo_col: "AJO"})
df = df.dropna(subset=["SFMC"])
df["SFMC"] = df["SFMC"].astype(str)
df["AJO"]  = df["AJO"].fillna("").astype(str)

# 1) Read HTML
html = Path(html_input).read_text(encoding="utf-8").replace("\r\n", "\n")

# 2) Known variables from the mapping
var_expr, covered_vars = extract_var_expressions(df)
warnings = []  # conditions will add unmapped vars here

# 3) Convert IF/ELSE/ENDIF + handle "the + PlanLegalName"
html_after_if = convert_ampscript_if_blocks(html, warnings, var_expr, covered_vars)

# 4) Table-driven literal replacements (de-para)
rows = sorted(df.to_dict("records"), key=lambda r: len(str(r["SFMC"]).strip()), reverse=True)
total_found = 0
total_replaced = 0
work = html_after_if

for r in rows:
    sfmc = str(r["SFMC"]).strip()
    ajo  = str(r["AJO"]).strip()
    if not sfmc:
        continue
    pattern = build_flex_regex(sfmc)
    matches = list(pattern.finditer(work))
    if not matches:
        continue
    total_found += len(matches)
    if ajo:
        work, n = pattern.subn(ajo, work)
        total_replaced += n

# 5) Global print replacement (only for mapped variables)
work = replace_all_prints(work, var_expr)

# 6) Comment remaining AMPScript safely + hoist any {% let ... %} inside
final_html, commented_blocks = comment_ampscript_with_hoist(work)

# 7) Save final HTML
Path(html_output).write_text(final_html, encoding="utf-8")

# 8) Excel log
log_xlsx = write_excel_log(
    commented_blocks=commented_blocks,
    total_found=total_found,
    total_replaced=total_replaced,
    output_path=html_output,
    warnings=warnings,
    covered_vars={v.lower() for v in covered_vars},
    var_expr=var_expr
)

# 9) Console summary
print("==== SFMC → AJO (report) ====")
print(f"Input   : {html_input}")
print(f"Output  : {html_output}")
print(f"Mapping : {excel_path}")
print(f"SFMC matches: {total_found}")
print(f"Substitutions: {total_replaced}")
print(f"Commented AMPScript: {len(commented_blocks)}")
print(f"Excel log: {log_xlsx.resolve()}")

# Extra heads-up for unmapped variables referenced in conditions
unresolved = sorted({w.lower() for w in warnings} - {v.lower() for v in covered_vars})
if unresolved:
    print("\nUnmapped variables detected:")
    for w in unresolved:
        print(f"  - @{w}  → set profile.{w} or context.{w} in AJO (or add to mapping sheet)")
